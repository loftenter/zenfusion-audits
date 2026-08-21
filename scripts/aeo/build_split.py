import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SPLIT=json.load(open("/home/claude/mhwf_deep/split.json"))
ENT=json.load(open("/home/claude/mhwf_deep/enriched.json"))

wb=Workbook()
H1=Font(name="Arial",bold=True,size=14,color="FFFFFF"); H2=Font(name="Arial",bold=True,size=10,color="FFFFFF")
LBL=Font(name="Arial",bold=True,size=10,color="173B5E"); TXT=Font(name="Arial",size=9)
TXTB=Font(name="Arial",size=10); LINK=Font(name="Arial",size=10,color="0563C1",underline="single")
NOTE=Font(name="Arial",italic=True,size=9,color="666666"); ENTF=Font(name="Arial",size=9,color="333333")
NAVY=PatternFill("solid",fgColor="173B5E"); GREEN=PatternFill("solid",fgColor="2E7D32")
ORANGE=PatternFill("solid",fgColor="C65911"); PURPLE=PatternFill("solid",fgColor="6A3D9A")
TEAL=PatternFill("solid",fgColor="2A8C8C"); LBLUE=PatternFill("solid",fgColor="D6E4F0")
AMBER=PatternFill("solid",fgColor="FFF2CC"); LORANGE=PatternFill("solid",fgColor="FCE4D6")
thin=Side(style="thin",color="D0D0D0"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
CTR=Alignment(horizontal="center",vertical="center"); LEFT=Alignment(horizontal="left",vertical="center")
WRAP=Alignment(wrap_text=True,vertical="top")
HOME="https://myhomewaterfilter.com"

def yoystr(y): return f"+{y}%" if isinstance(y,int) and y>0 else (f"{y}%" if isinstance(y,int) else "")

order=sorted(SPLIT.keys(), key=lambda c:-sum((r[1] or 0) for r in SPLIT[c]["generic"]))
def tabname(i,c):
    short=c.replace(" / ","-").replace("/","-").replace(" & "," ")
    return f"#{i} {short}"[:31]

# README
ws=wb.active; ws.title="📋 README"; ws.sheet_view.showGridLines=False
ws.column_dimensions['A'].width=3; ws.column_dimensions['B'].width=30; ws.column_dimensions['C'].width=104
ws['B2']="MyHomeWaterFilter.com — DEEP Keyword + Entity Map (Generic vs Branded split)"
ws['B2'].font=Font(name="Arial",bold=True,size=15,color="173B5E")
ws['B3']="Live DataforSEO keyword_suggestions across all cores · 6,477 unique US keywords · each tab = full secondary depth, split into GENERIC (unbranded) and BRANDED (competitor/retailer) blocks side by side, plus a derived entity panel."
ws['B3'].font=NOTE
rows=[
("Per-tab layout","THREE blocks, left to right: (1) GENERIC KEYWORDS [cols B–F] — category terms with NO brand/product/retailer names. (2) BRANDED KEYWORDS [cols H–L] — every term containing a brand (SoftPro, WaterDrop, or any competitor) or retailer (Costco, Home Depot, Amazon, Lowe's). (3) ENTITY & SEMANTIC PANEL [cols N–Q]."),
("Why the split","The generic block is your pure category-demand target list (what you rank/build pages for). The branded block shows competitor & retailer demand with traffic — intel for comparison pages, conquesting, and seeing where shoppers attach a brand — kept separate so it never dilutes the generic list."),
("Both blocks","Keyword · Volume · KD · Intent · YoY, sorted by volume. Amber = KD ≤ 20 quick win (generic) / FCE4D6 peach header marks the branded block."),
("Branded = any brand","A term is 'branded' if it contains ANY brand/product/manufacturer/retailer name — including your own (SoftPro, WaterDrop) and competitors (SpringWell, Aquasana, Brita, Berkey, Culligan, etc.) and stores (Costco, Home Depot, Lowe's, Amazon, Walmart)."),
("Entity columns","Synonyms & NLP Terms = extracted from the cluster's real corpus. Google Entities = brand/org/standard/material/contaminant/product-type nodes co-occurring in corpus. AI Entities = associated concepts DERIVED from corpus+topic (no live LLM API on this plan)."),
("Notes","Blank KD = endpoint didn't return it this run (errors intermittently), not zero. Volumes are highest-month-normalized; some carry 2025 PFAS-spike inflation — use 12-mo judgment. Combined volumes count overlapping phrasings (relative scale, not additive)."),
]
r=5
for lab,val in rows:
    ws.cell(row=r,column=2,value=lab).font=LBL; ws.cell(row=r,column=2).alignment=Alignment(vertical="top")
    c=ws.cell(row=r,column=3,value=val); c.font=TXT; c.alignment=WRAP; ws.row_dimensions[r].height=58; r+=1
r+=1
ws.cell(row=r,column=2,value="CLUSTER INDEX (by generic volume)").font=LBL; r+=1
for i,c in enumerate(order,1):
    g=SPLIT[c]["generic"]; b=SPLIT[c]["branded"]; gv=sum((x[1] or 0) for x in g)
    ws.cell(row=r,column=2,value=tabname(i,c)).font=TXTB
    ws.cell(row=r,column=3,value=f"{len(g)} generic · {len(b)} branded · ~{gv:,}/mo generic · core: {c}").font=TXT
    r+=1

def write_kw_block(ws, startcol, header_fill, title, kws, amber=True):
    # title bar across 5 cols
    ws.merge_cells(start_row=4,start_column=startcol,end_row=4,end_column=startcol+4)
    tc=ws.cell(row=4,column=startcol,value=title); tc.font=H2; 
    for cc in range(startcol,startcol+5): ws.cell(row=4,column=cc).fill=header_fill
    # header row 5
    for j,h in enumerate(["Keyword","Vol","KD","Intent","YoY"]):
        cell=ws.cell(row=5,column=startcol+j,value=h)
        cell.font=Font(name="Arial",bold=True,size=9,color="FFFFFF"); cell.fill=header_fill
        cell.alignment=LEFT if j==0 else CTR; cell.border=BORDER
    rr=6
    for kw,vol,kd,intent,yoy,cpc in kws:
        ws.cell(row=rr,column=startcol,value=kw).font=TXT; ws.cell(row=rr,column=startcol).border=BORDER
        ws.cell(row=rr,column=startcol+1,value=vol if vol else "").font=TXT; ws.cell(row=rr,column=startcol+1).alignment=CTR; ws.cell(row=rr,column=startcol+1).border=BORDER
        ws.cell(row=rr,column=startcol+2,value=kd if kd is not None else "").font=TXT; ws.cell(row=rr,column=startcol+2).alignment=CTR; ws.cell(row=rr,column=startcol+2).border=BORDER
        ws.cell(row=rr,column=startcol+3,value=intent or "").font=TXT; ws.cell(row=rr,column=startcol+3).alignment=CTR; ws.cell(row=rr,column=startcol+3).border=BORDER
        ws.cell(row=rr,column=startcol+4,value=yoystr(yoy)).font=TXT; ws.cell(row=rr,column=startcol+4).alignment=CTR; ws.cell(row=rr,column=startcol+4).border=BORDER
        if amber and isinstance(kd,int) and kd<=20:
            for cc in range(startcol,startcol+5): ws.cell(row=rr,column=cc).fill=AMBER
        rr+=1
    return rr

for i,c in enumerate(order,1):
    gen=SPLIT[c]["generic"]; brd=SPLIT[c]["branded"]; d=ENT[c]
    gv=sum((x[1] or 0) for x in gen); bv=sum((x[1] or 0) for x in brd)
    ws=wb.create_sheet(tabname(i,c)); ws.sheet_view.showGridLines=False
    widths=[('A',2),('B',44),('C',10),('D',6),('E',12),('F',8),('G',2),
            ('H',44),('I',10),('J',6),('K',12),('L',8),('M',2),
            ('N',26),('O',22),('P',26),('Q',26)]
    for col,w in widths: ws.column_dimensions[col].width=w
    # title
    ws.merge_cells('B2:L2'); t=ws['B2']; t.value=f"CORE KEYWORD: {c.upper()}"; t.font=H1
    for col in range(2,13): ws.cell(row=2,column=col).fill=NAVY
    ws.row_dimensions[2].height=24
    ws.merge_cells('B3:L3'); s=ws['B3']
    s.value=f"GENERIC: {len(gen)} kw (~{gv:,}/mo)   |   BRANDED: {len(brd)} kw (~{bv:,}/mo)   |   source: DataforSEO keyword_suggestions (US)"
    s.font=Font(name="Arial",bold=True,size=10,color="173B5E")
    for col in range(2,13): ws.cell(row=3,column=col).fill=LBLUE

    r1=write_kw_block(ws,2,GREEN,"GENERIC KEYWORDS (unbranded category demand)",gen,amber=True)
    r2=write_kw_block(ws,8,ORANGE,"BRANDED KEYWORDS (competitor / retailer demand)",brd,amber=False)
    # peach tint the branded header row label area already orange; keep

    # entity panel cols N-Q (14-17)
    ws.merge_cells(start_row=4,start_column=14,end_row=4,end_column=17)
    eh=ws.cell(row=4,column=14,value="ENTITY & SEMANTIC PANEL (derived)"); eh.font=H2
    for col in range(14,18): ws.cell(row=4,column=col).fill=PURPLE
    for j,title in enumerate(["Synonyms","NLP Terms","Google Entities","AI Entities (derived)"]):
        cell=ws.cell(row=5,column=14+j,value=title); cell.font=Font(name="Arial",bold=True,size=9,color="FFFFFF")
        cell.fill=TEAL if j<2 else PURPLE; cell.alignment=CTR; cell.border=BORDER
    cols_data=[d["synonyms"],d["nlp"],d["google_entities"],d["ai_entities"]]
    maxlen=max((len(x) for x in cols_data), default=0)
    for k in range(maxlen):
        for j,lst in enumerate(cols_data):
            val=lst[k] if k<len(lst) else ""
            cell=ws.cell(row=6+k,column=14+j,value=val); cell.font=ENTF; cell.alignment=WRAP; cell.border=BORDER

    ws.freeze_panes="B6"

out="/home/claude/mhwf_deep/MyHomeWaterFilter_DEEP_Generic_vs_Branded.xlsx"
wb.save(out); print("Saved",out,"tabs",len(wb.worksheets))
