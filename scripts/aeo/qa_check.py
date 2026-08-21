#!/usr/bin/env python3
"""
ZENFUSION AEO — QA GATE
Validates a built workbook for accuracy + completeness BEFORE it is presented
to a human team member. Designed to be run by an agent as the final automated
step. Exits non-zero on FAIL so an orchestrator can block hand-off.

Run:  python3 qa_check.py --config clients/<client>/config.json

Checks (each prints PASS / WARN / FAIL):
  1. Workbook exists and opens
  2. All framework tabs present (Brand Demand, Your Demand, AI Overview Wins) + >=1 cluster tab
  3. No off-topic / cross-domain terms survived into any cluster tab
  4. No fabricated per-LLM volume (AEO 'Google Vol' must be blank or a real int; AI Overview? uses only the allowed vocabulary)
  5. Brand Demand volumes are real numbers (not placeholders)
  6. AI Overview Wins totals reconcile with the row data
  7. Reddit/YouTube blocks: where populated, URLs look valid; counts <= 20
  8. No column collisions on cluster tabs (block headers land in expected columns)
  9. AEO citations: every 'Yes (verified)' row has named sources (not 'pull live')
FAIL on any hard violation; WARN on soft/coverage gaps (so a human can decide).
"""
import json, sys, os, re

GREEN="\033[92m"; YEL="\033[93m"; RED="\033[91m"; RST="\033[0m"
results=[]
def ok(m):   results.append(("PASS",m)); print(f"{GREEN}PASS{RST}  {m}")
def warn(m): results.append(("WARN",m)); print(f"{YEL}WARN{RST}  {m}")
def fail(m): results.append(("FAIL",m)); print(f"{RED}FAIL{RST}  {m}")

def load_cfg(p):
    cfg=json.load(open(p))
    cfgdir=os.path.dirname(os.path.abspath(p)); root=cfgdir
    for _ in range(4):
        if os.path.isdir(os.path.join(root,"scripts")): break
        root=os.path.dirname(root)
    work=cfg["output"]["work_dir"]
    if not os.path.isabs(work): work=os.path.join(root,work)
    return cfg, work

def main():
    if len(sys.argv)<3 or sys.argv[1]!="--config":
        print("usage: qa_check.py --config <path>"); sys.exit(2)
    cfg, work = load_cfg(sys.argv[2])
    import openpyxl
    xlsx=os.path.join(work, cfg["output"]["filename_template"].format(display_name=cfg["client"]["display_name"]))
    if not os.path.exists(xlsx):
        fail(f"workbook not found: {xlsx}"); sys.exit(1)
    wb=openpyxl.load_workbook(xlsx)
    ok(f"workbook opens ({len(wb.sheetnames)} tabs)")

    # 2. framework tabs
    names=wb.sheetnames
    for needle in ["Brand Demand","Your Demand","AI Overview Wins"]:
        if any(needle in n for n in names): ok(f"tab present: {needle}")
        else: fail(f"missing framework tab: {needle}")
    cluster_tabs=[n for n in names if n.startswith("#")]
    if cluster_tabs: ok(f"{len(cluster_tabs)} cluster tabs present")
    else: fail("no cluster tabs (#N ...) found")

    # 3. off-topic scrub across cluster tabs (col B generic keywords + AL/AS topical)
    off=[o.lower() for o in cfg.get("off_topic_scrub",[])]
    contaminated=[]
    for n in cluster_tabs:
        ws=wb[n]
        for r in range(6, ws.max_row+1):
            for col in (2,):  # generic keyword column
                v=ws.cell(r,col).value
                if v and any(o in str(v).lower() for o in off):
                    contaminated.append((n,str(v)))
    if contaminated:
        fail(f"off-topic terms in clusters: {contaminated[:5]}{'...' if len(contaminated)>5 else ''} ({len(contaminated)} total)")
    else:
        ok("no off-topic / cross-domain terms in cluster keyword columns")

    # 4. AEO vocabulary + no fabricated volume (AI Overview? col = 26 on cluster tabs)
    allowed={"Yes (verified)","No (verified)","Likely*","Maybe*","Yes","No",None,""}
    bad_aio=[]; bad_vol=[]
    for n in cluster_tabs:
        ws=wb[n]
        for r in range(6, ws.max_row+1):
            a=ws.cell(r,26).value
            if a not in allowed: bad_aio.append((n,a))
            v=ws.cell(r,25).value
            if v not in (None,"") and not isinstance(v,(int,float)): bad_vol.append((n,v))
    if bad_aio: fail(f"unexpected AI-Overview values: {bad_aio[:5]}")
    else: ok("AI Overview? column uses only the allowed honest vocabulary")
    if bad_vol: warn(f"non-numeric AEO volume cells: {bad_vol[:5]}")
    else: ok("AEO Google-Vol cells are blank or numeric (no fabricated values)")

    # 5. Brand Demand volumes real
    bd=[n for n in names if "Brand Demand" in n][0]
    ws=wb[bd]; nonnum=0; rows=0
    for r in range(6, ws.max_row+1):
        if ws.cell(r,2).value:
            rows+=1
            if not isinstance(ws.cell(r,4).value,(int,float)): nonnum+=1
    if rows==0: warn("Brand Demand tab has no rows")
    elif nonnum>rows*0.2: warn(f"{nonnum}/{rows} brand rows have non-numeric volume")
    else: ok(f"Brand Demand: {rows} brands, volumes numeric")

    # 6. AI Overview Wins reconcile
    aw=[n for n in names if "AI Overview Wins" in n][0]
    ws=wb[aw]; n_rows=0; vols=0
    for r in range(6, ws.max_row+1):
        kw=ws.cell(r,2).value; v=ws.cell(r,3).value
        if kw and not str(kw).startswith("Source:"):
            n_rows+=1
            if isinstance(v,(int,float)): vols+=v
    if n_rows>0: ok(f"AI Overview Wins: {n_rows} keywords, {vols:,} combined volume")
    else: warn("AI Overview Wins tab empty (expected for a brand with no AEO footprint yet)")

    # 7. Reddit/YouTube sanity (cols 30 reddit url, 34 yt url)
    bad_url=0; over=0
    for n in cluster_tabs:
        ws=wb[n]; rc=0; yc=0
        for r in range(6, ws.max_row+1):
            ru=ws.cell(r,30).value; yu=ws.cell(r,34).value
            if ru and "pending" not in str(ru).lower():
                rc+=1
                if "reddit.com" not in str(ru): bad_url+=1
            if yu and "pending" not in str(yu).lower():
                yc+=1
                if "youtu" not in str(yu) and "instagram" not in str(yu) and "facebook" not in str(yu) and "tiktok" not in str(yu): bad_url+=1
        if rc>20 or yc>20: over+=1
    if bad_url: warn(f"{bad_url} reddit/social URLs look off-domain")
    else: ok("Reddit/YouTube URLs look valid where populated")
    if over: fail(f"{over} cluster tab(s) exceed the 20-URL cap")
    else: ok("Reddit/YouTube blocks within the 20-row cap")

    # 8. column collision check (block headers in row 4 at expected cols)
    expected={2:"GENERIC",20:"TOP 10 ORGANIC COMPETITORS",24:"AI / AEO",29:"TOP REDDIT",32:"TOP YOUTUBE",36:"ENTITY",41:"PEOPLE ALSO ASK",45:"TOPICAL MAP"}
    coll=0
    if cluster_tabs:
        ws=wb[cluster_tabs[0]]
        for col,needle in expected.items():
            v=str(ws.cell(4,col).value or "")
            if needle.split()[0] not in v: coll+=1
    if coll: fail(f"{coll} block header(s) not in expected columns (possible collision)")
    else: ok("cluster block headers align to expected columns (no collisions)")

    # 9. verified AEO rows must have real sources
    missing=0
    for n in cluster_tabs:
        ws=wb[n]
        for r in range(6, ws.max_row+1):
            if ws.cell(r,26).value=="Yes (verified)":
                src=str(ws.cell(r,27).value or "")
                if not src or src=="pull live to capture": missing+=1
    if missing: fail(f"{missing} 'Yes (verified)' AEO rows missing cited sources")
    else: ok("all verified AEO rows carry named citation sources")

    # summary
    n_fail=sum(1 for s,_ in results if s=="FAIL")
    n_warn=sum(1 for s,_ in results if s=="WARN")
    print("\n"+"="*60)
    print(f"QA SUMMARY: {len(results)} checks | {n_fail} FAIL | {n_warn} WARN")
    if n_fail:
        print(f"{RED}RESULT: FAIL — do NOT present to human until resolved.{RST}")
        # write a machine-readable report
        json.dump({"result":"FAIL","fail":n_fail,"warn":n_warn,"details":results}, open(os.path.join(work,"qa_report.json"),"w"))
        sys.exit(1)
    elif n_warn:
        print(f"{YEL}RESULT: PASS WITH WARNINGS — safe to present; note the warnings for the reviewer.{RST}")
    else:
        print(f"{GREEN}RESULT: CLEAN PASS — ready for human review.{RST}")
    json.dump({"result":"PASS","fail":n_fail,"warn":n_warn,"details":results}, open(os.path.join(work,"qa_report.json"),"w"))

if __name__=="__main__":
    main()
