import json, sys, re, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================================
# DUAL ENTRY MODE
#   Config mode (preferred):  python build_paa.py --config path/to/client_config.json
#   Legacy positional mode :  python build_paa.py split ent scriptdir home out prop paa comp aeo ryt aeolive brands yourdemand aeowins
# In config mode, all data files are resolved from <work_dir>/data/*.json by
# convention, so a new client is one config file + a populated data/ folder.
# ============================================================================
def _load(path, default):
    try:
        return json.load(open(path))
    except (FileNotFoundError, json.JSONDecodeError):
        return default

if len(sys.argv) >= 3 and sys.argv[1] == "--config":
    CFG = json.load(open(sys.argv[2]))
    CFG_DIR = os.path.dirname(os.path.abspath(sys.argv[2]))
    # package root = the dir that contains 'scripts/'. Walk up from the config file to find it.
    pkg_root = CFG_DIR
    for _ in range(4):
        if os.path.isdir(os.path.join(pkg_root, "scripts")):
            break
        pkg_root = os.path.dirname(pkg_root)
    WORK = CFG["output"]["work_dir"]
    if not os.path.isabs(WORK):
        WORK = os.path.join(pkg_root, WORK)
    DATA = os.path.join(WORK, "data")
    sys.path.insert(0, os.path.join(pkg_root, "scripts"))
    SPLIT     = _load(os.path.join(DATA, "split.json"), {})
    ENT       = _load(os.path.join(DATA, "enriched.json"), {})
    PAA       = _load(os.path.join(DATA, "paa_all.json"), {})
    COMP      = _load(os.path.join(DATA, "competitors.json"), {})
    AEO       = _load(os.path.join(DATA, "aeo.json"), {})
    RYT       = _load(os.path.join(DATA, "reddit_yt.json"), {})
    AEOLIVE   = _load(os.path.join(DATA, "aeo_live.json"), {})
    BRANDS    = _load(os.path.join(DATA, "brands.json"), [])
    YOURDEMAND= _load(os.path.join(DATA, "yourdemand.json"), {})
    AEOWINS   = _load(os.path.join(DATA, "aeo_wins.json"), [])
    PROP      = CFG["client"]["display_name"]
    HOME      = CFG["client"].get("url", "")
    OUT       = os.path.join(WORK, CFG["output"]["filename_template"].format(display_name=PROP))
    _PALETTE  = CFG.get("brand_palette", {})
else:
    sys.path.insert(0, sys.argv[3])  # dir containing segment_engine
    SPLIT=json.load(open(sys.argv[1]))
    ENT=json.load(open(sys.argv[2]))
    HOME=sys.argv[4]; OUT=sys.argv[5]; PROP=sys.argv[6]
    PAA=json.load(open(sys.argv[7])) if len(sys.argv)>7 else {}
    COMP=json.load(open(sys.argv[8])) if len(sys.argv)>8 else {}
    AEO=json.load(open(sys.argv[9])) if len(sys.argv)>9 else {}
    RYT=json.load(open(sys.argv[10])) if len(sys.argv)>10 else {}
    AEOLIVE=json.load(open(sys.argv[11])) if len(sys.argv)>11 else {}
    BRANDS=json.load(open(sys.argv[12])) if len(sys.argv)>12 else []
    YOURDEMAND=json.load(open(sys.argv[13])) if len(sys.argv)>13 else {}
    AEOWINS=json.load(open(sys.argv[14])) if len(sys.argv)>14 else []
    _PALETTE={}

from segment_engine import tag_keyword, detect_property_segments, FAMILIES
from topical_engine import build_topical_map

# Map a cluster/tab label to a clean, natural core phrase for the topical map.
def core_phrase(cluster):
    AWG = "generator" in PROP.lower() or "watergeneratorpros" in PROP.lower()
    if AWG:
        AOVR={"Home / Residential":"Home Atmospheric Water Generator",
              "Best / Reviews / Compare":"Best Atmospheric Water Generator",
              "Price / Cost / Buy":"Atmospheric Water Generator Cost",
              "Portable / Off-Grid / Solar":"Portable Atmospheric Water Generator",
              "DIY / How-To / Informational":"DIY Atmospheric Water Generator",
              "Commercial / Industrial / Ag":"Commercial Atmospheric Water Generator",
              "Water From Air / Air-to-Water (generic)":"Water From Air Machine",
              "Water Generator (generic)":"Water Generator","AWG Machine":"AWG Machine",
              "Brand: Watergen":"Watergen Atmospheric Water Generator",
              "Competitor Brands":"Atmospheric Water Generator Brands",
              "Atmospheric Water Generator (head)":"Atmospheric Water Generator"}
        if cluster in AOVR: return AOVR[cluster]
    OVR={
     "Whole House / POE":"Whole House Water Filter","Faucet & Inline":"Faucet Water Filter",
     "Home / Residential":"Home Water Filter","Countertop & Pitcher":"Countertop Water Filter",
     "Under-Sink":"Under-Sink Water Filter","Refrigerator/Fridge":"Refrigerator Water Filter",
     "Carbon / Media":"Carbon Water Filter","PFAS / Lead":"PFAS Water Filter",
     "UV / Disinfection":"UV Water Purifier","Sediment & Specialty Well":"Sediment Filter",
     "Alkaline / Ionizer / Hydrogen":"Alkaline Water Filter","Chlorine / Chloramine":"Chlorine Filter",
     "Best / Reviews / Compare":"Best Water Filter","Brand & Competitor":"Water Filter Brands",
     "Brand: Watergen":"Watergen Atmospheric Water Generator","Competitor Brands":"Atmospheric Water Generator Brands",
     "Best / Reviews / Compare ":"Best Atmospheric Water Generator","AWG Machine":"AWG Machine",
     "Water From Air / Air-to-Water (generic)":"Water From Air Machine","Water Generator (generic)":"Water Generator",
     "DIY / How-To / Informational":"DIY Atmospheric Water Generator","Commercial / Industrial / Ag":"Commercial Atmospheric Water Generator",
     "Price / Cost / Buy":"Atmospheric Water Generator","Portable / Off-Grid / Solar":"Portable Atmospheric Water Generator",
    }
    if cluster in OVR: return OVR[cluster]
    import re as _re
    s=cluster
    # strip parentheticals and slashes -> take the first, most representative noun phrase
    s=_re.sub(r"\(.*?\)","",s)
    s=s.split("/")[0].split("&")[0].split(" vs ")[0]
    s=s.replace(" - "," ").strip()
    # title-case-ish but keep acronyms
    fixes={"Poe":"POE","Ro":"RO","Uv":"UV","Pfas":"PFAS","Awg":"AWG","Genny":"GENNY","Diy":"DIY"}
    words=[]
    for w in s.split():
        wl=w.capitalize() if not w.isupper() else w
        words.append(fixes.get(wl,w if w.isupper() else wl))
    out=" ".join(words).strip()
    return out or cluster

wb=Workbook()
# ===== Zenfusion brand palette =====
Z_GREEN=_PALETTE.get("green","35EEA0"); Z_BLUE=_PALETTE.get("blue","30C8EE"); Z_NAVY=_PALETTE.get("navy","041952"); Z_CHAR=_PALETTE.get("charcoal","1A1A1A")
Z_GREEN_LT="D7FBEC"; Z_BLUE_LT="D6F4FB"; Z_NAVY_LT="DDE3F0"; Z_BAND="EAF6F1"
H1=Font(name="Arial",bold=True,size=14,color="FFFFFF")
H2=Font(name="Arial",bold=True,size=10,color="FFFFFF")
LBL=Font(name="Arial",bold=True,size=10,color="041952")
TXT=Font(name="Arial",size=9)
TXTB=Font(name="Arial",size=10); LINK=Font(name="Arial",size=10,color="0563C1",underline="single")
NOTE=Font(name="Arial",italic=True,size=9,color="666666"); ENTF=Font(name="Arial",size=9,color="333333")
NAVY=PatternFill("solid",fgColor=Z_NAVY); GREEN=PatternFill("solid",fgColor=Z_NAVY)
ORANGE=PatternFill("solid",fgColor=Z_CHAR); BLUESEG=PatternFill("solid",fgColor=Z_BLUE)
PURPLE=PatternFill("solid",fgColor=Z_NAVY); TEAL=PatternFill("solid",fgColor=Z_BLUE)
PAAFILL=PatternFill("solid",fgColor=Z_CHAR); PAALITE=PatternFill("solid",fgColor=Z_BAND)
TOPHEAD=PatternFill("solid",fgColor=Z_NAVY); TOPALT=PatternFill("solid",fgColor=Z_BAND)
LBLUE=PatternFill("solid",fgColor=Z_NAVY_LT); AMBER=PatternFill("solid",fgColor=Z_GREEN_LT)
GREENHDR=PatternFill("solid",fgColor=Z_GREEN)  # elevated Shopify green for headers
AEOHDR=PatternFill("solid",fgColor=Z_BLUE)     # bright blue for AI/AEO block
REDDITHDR=PatternFill("solid",fgColor=Z_CHAR)  # charcoal for Reddit block
YTHDR=PatternFill("solid",fgColor=Z_NAVY)      # navy for YouTube block
thin=Side(style="thin",color="D0D0D0"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
CTR=Alignment(horizontal="center",vertical="center"); LEFT=Alignment(horizontal="left",vertical="center")
WRAP=Alignment(wrap_text=True,vertical="top")

def yoystr(y): return f"+{y}%" if isinstance(y,int) and y>0 else (f"{y}%" if isinstance(y,int) else "")
order=sorted(SPLIT.keys(), key=lambda c:-sum((r[1] or 0) for r in SPLIT[c]["generic"]))
def tabname(i,c):
    short=c.replace(":","").replace(" / ","-").replace("/","-").replace(" & "," ")
    return f"#{i} {short}"[:31]

# property-level detected segments (for README + legend)
allk=[]
for c in SPLIT: allk+=SPLIT[c]["generic"]
prop_segs=detect_property_segments(allk)
seg_legend=", ".join(n for n,_ in prop_segs[:14])

# README
ws=wb.active; ws.title="① Brand Demand"; ws.sheet_view.showGridLines=False
ws.column_dimensions['A'].width=2
for col,w in [('B',34),('C',22),('D',14),('E',12),('F',12),('G',14)]:
    ws.column_dimensions[col].width=w
ws.merge_cells('B2:G2'); t=ws['B2']; t.value=f"{PROP} — BRAND & PRODUCT SEARCH DEMAND"; t.font=H1
for col in range(2,8): ws.cell(row=2,column=col).fill=NAVY
ws.row_dimensions[2].height=24
ws.merge_cells('B3:G3'); s=ws['B3']
s.value="What people search for by brand/product name, per month (Google US, live DataforSEO). Ranked by volume. Use this to see where branded demand actually sits — yours vs competitors."
s.font=NOTE
# header row
hdr=["Brand / Product","Type","Monthly Vol","CPC ($)","Competition","Trend (YoY)"]
for j,h in enumerate(hdr):
    cell=ws.cell(row=5,column=2+j,value=h); cell.font=Font(name="Arial",bold=True,size=10,color="041952")
    cell.fill=GREENHDR; cell.alignment=LEFT if j in (0,1) else CTR; cell.border=BORDER
rr=6
for b in BRANDS:
    is_soft = "SoftPro" in b.get("type","") or b["brand"].lower().startswith("softpro")
    ws.cell(row=rr,column=2,value=b["brand"]).font=Font(name="Arial",bold=is_soft,size=9,color="041952" if is_soft else "000000")
    ws.cell(row=rr,column=2).border=BORDER; ws.cell(row=rr,column=2).alignment=LEFT
    ws.cell(row=rr,column=3,value=b["type"]).font=TXT; ws.cell(row=rr,column=3).border=BORDER; ws.cell(row=rr,column=3).alignment=LEFT
    ws.cell(row=rr,column=4,value=b["volume"]).font=TXT; ws.cell(row=rr,column=4).number_format="#,##0"; ws.cell(row=rr,column=4).alignment=CTR; ws.cell(row=rr,column=4).border=BORDER
    ws.cell(row=rr,column=5,value=b["cpc"]).font=TXT; ws.cell(row=rr,column=5).number_format='$0.00'; ws.cell(row=rr,column=5).alignment=CTR; ws.cell(row=rr,column=5).border=BORDER
    ws.cell(row=rr,column=6,value=b["competition"]).font=TXT; ws.cell(row=rr,column=6).alignment=CTR; ws.cell(row=rr,column=6).border=BORDER
    ws.cell(row=rr,column=7,value=yoystr(b["yoy"])).font=TXT; ws.cell(row=rr,column=7).alignment=CTR; ws.cell(row=rr,column=7).border=BORDER
    if is_soft:
        for col in range(2,8): ws.cell(row=rr,column=col).fill=PatternFill("solid",fgColor=Z_GREEN_LT)
    rr+=1
# footnote
fr=rr+1
ws.merge_cells(start_row=fr,start_column=2,end_row=fr,end_column=7)
fn=ws.cell(row=fr,column=2,value="Monthly Vol = Google US exact-match search volume (last reported month). CPC = top-of-page cost-per-click. Competition = Google Ads index 0–100. Trend (YoY) = recent 3-mo avg vs trailing 3-mo avg of the 12-mo window. SoftPro rows highlighted. Blank/omitted brands = volume too low for Google to report.")
fn.font=NOTE; fn.alignment=WRAP
ws.freeze_panes="B6"
if not BRANDS:
    ws.cell(row=6,column=2,value="(brand data pending)").font=NOTE

# ===== TAB 2: YOUR DEMAND (brand-owner's own branded keywords for THIS site) =====
wsd=wb.create_sheet("② Your Demand"); wsd.sheet_view.showGridLines=False
wsd.column_dimensions['A'].width=2
for col,w in [('B',40),('C',22),('D',14),('E',12),('F',14),('G',14)]:
    wsd.column_dimensions[col].width=w
bn = YOURDEMAND.get("brand_name", PROP)
wsd.merge_cells('B2:G2'); td=wsd['B2']; td.value=f"YOUR DEMAND — branded search for {bn}"; td.font=H1
for col in range(2,8): wsd.cell(row=2,column=col).fill=NAVY
wsd.row_dimensions[2].height=24
wsd.merge_cells('B3:G3'); sd=wsd['B3']
sd.value="Search demand for YOUR OWN brand name, domain, and brand+modifier phrases (Google US, live). This is your branded-search footprint — how many people look for you by name each month. New brands often start at zero; this tab is the baseline you grow over time."
sd.font=NOTE
hdr=["Your Brand Keyword / Phrase","Type","Monthly Vol","CPC ($)","Competition","Trend (YoY)"]
for j,h in enumerate(hdr):
    cell=wsd.cell(row=5,column=2+j,value=h); cell.font=Font(name="Arial",bold=True,size=10,color="041952")
    cell.fill=GREENHDR; cell.alignment=LEFT if j in (0,1) else CTR; cell.border=BORDER
rr=6
checked = YOURDEMAND.get("checked", [])
for row in checked:
    kw,typ,vol,cpc,comp = (row + [None,None,None])[:5]
    wsd.cell(row=rr,column=2,value=kw).font=Font(name="Arial",bold=True,size=9,color="041952"); wsd.cell(row=rr,column=2).border=BORDER; wsd.cell(row=rr,column=2).alignment=LEFT
    wsd.cell(row=rr,column=3,value=typ).font=TXT; wsd.cell(row=rr,column=3).border=BORDER; wsd.cell(row=rr,column=3).alignment=LEFT
    wsd.cell(row=rr,column=4,value=(vol if vol else "—")).font=TXT; wsd.cell(row=rr,column=4).alignment=CTR; wsd.cell(row=rr,column=4).border=BORDER
    if vol: wsd.cell(row=rr,column=4).number_format="#,##0"
    wsd.cell(row=rr,column=5,value=(cpc if cpc else "—")).font=TXT; wsd.cell(row=rr,column=5).alignment=CTR; wsd.cell(row=rr,column=5).border=BORDER
    if cpc: wsd.cell(row=rr,column=5).number_format='$0.00'
    wsd.cell(row=rr,column=6,value=(comp if comp is not None else "—")).font=TXT; wsd.cell(row=rr,column=6).alignment=CTR; wsd.cell(row=rr,column=6).border=BORDER
    wsd.cell(row=rr,column=7,value="—").font=TXT; wsd.cell(row=rr,column=7).alignment=CTR; wsd.cell(row=rr,column=7).border=BORDER
    rr+=1
# baseline note
total_vol = sum((r[2] or 0) for r in checked)
fr=rr+1
wsd.merge_cells(start_row=fr,start_column=2,end_row=fr,end_column=7)
if total_vol==0:
    msg=("BASELINE: No measurable branded search volume yet — expected for a newly launched brand. "
         "As brand awareness grows (PR, social, AI-citation wins, word of mouth), branded searches appear here first. "
         "Re-run quarterly to track your branded-demand curve. A rising 'Your Demand' total is one of the clearest signals that top-of-funnel brand building is working. '—' = no data returned (volume below Google's reporting threshold).")
else:
    msg=("Branded-search footprint for your own brand. Re-run quarterly to track the trend. "
         "Rising branded demand is a leading indicator that brand-building is working. '—' = below Google's reporting threshold.")
fn=wsd.cell(row=fr,column=2,value=msg); fn.font=NOTE; fn.alignment=WRAP
wsd.row_dimensions[fr].height=72
wsd.freeze_panes="B6"

# ===== TAB 3: AI OVERVIEW WINS (keywords where the client's domain is cited in Google AI Overviews) =====
if AEOWINS:
    wsa=wb.create_sheet("③ AI Overview Wins"); wsa.sheet_view.showGridLines=False
    wsa.column_dimensions['A'].width=2
    for col,w in [('B',52),('C',14),('D',10),('E',40)]:
        wsa.column_dimensions[col].width=w
    wsa.merge_cells('B2:E2'); ta=wsa['B2']; ta.value="AI OVERVIEW WINS — where your brand is already cited by Google's AI"; ta.font=H1
    for col in range(2,6): wsa.cell(row=2,column=col).fill=NAVY
    wsa.row_dimensions[2].height=24
    tot_vol=sum(r[1] for r in AEOWINS); pos1=sum(1 for r in AEOWINS if r[2]==1)
    wsa.merge_cells('B3:E3'); sa=wsa['B3']
    sa.value=(f"Live from DataforSEO: {len(AEOWINS):,} keywords where the client's domains are cited as a source inside Google's AI Overview — {tot_vol:,} combined monthly searches, {pos1:,} at citation position #1. This is the strongest AEO asset in the report: proof the brand is already feeding Google's AI answers. Defend and expand these.")
    sa.font=NOTE
    wsa.row_dimensions[3].height=42
    hdr=["Keyword (AI Overview cites you)","Monthly Vol","Cite Pos","Cited Domain(s)"]
    for j,h in enumerate(hdr):
        cell=wsa.cell(row=5,column=2+j,value=h); cell.font=Font(name="Arial",bold=True,size=10,color="041952")
        cell.fill=GREENHDR; cell.alignment=LEFT if j in (0,3) else CTR; cell.border=BORDER
    rr=6
    for kw,vol,pos,dom in AEOWINS:
        wsa.cell(row=rr,column=2,value=kw).font=TXT; wsa.cell(row=rr,column=2).border=BORDER; wsa.cell(row=rr,column=2).alignment=LEFT
        wsa.cell(row=rr,column=3,value=vol).font=TXT; wsa.cell(row=rr,column=3).number_format="#,##0"; wsa.cell(row=rr,column=3).alignment=CTR; wsa.cell(row=rr,column=3).border=BORDER
        wsa.cell(row=rr,column=4,value=pos).font=TXT; wsa.cell(row=rr,column=4).alignment=CTR; wsa.cell(row=rr,column=4).border=BORDER
        wsa.cell(row=rr,column=5,value=dom).font=TXT; wsa.cell(row=rr,column=5).border=BORDER; wsa.cell(row=rr,column=5).alignment=LEFT
        if pos==1:
            wsa.cell(row=rr,column=4).fill=PatternFill("solid",fgColor=Z_GREEN_LT)
        rr+=1
    fr=rr+1
    wsa.merge_cells(start_row=fr,start_column=2,end_row=fr,end_column=5)
    fn=wsa.cell(row=fr,column=2,value="Source: DataforSEO ranked_keywords, item_type = ai_overview_reference (live). 'Cite Pos' = the brand's citation slot within the AI Overview. Position-1 rows are highlighted. This list is deduped across the client's domains; re-run quarterly to track AEO share-of-voice.")
    fn.font=NOTE; fn.alignment=WRAP; wsa.row_dimensions[fr].height=42
    wsa.freeze_panes="B6"

rows=[]  # README content removed; brand tabs are now tabs 1-2

def kw_block(ws, startcol, fill, title, kws, amber=True):
    bright = fill.fgColor.rgb in (Z_GREEN, Z_BLUE, "00"+Z_GREEN, "00"+Z_BLUE)
    htext = "041952" if bright else "FFFFFF"
    ws.merge_cells(start_row=4,start_column=startcol,end_row=4,end_column=startcol+4)
    tc=ws.cell(row=4,column=startcol,value=title); tc.font=Font(name="Arial",bold=True,size=10,color=htext)
    for cc in range(startcol,startcol+5): ws.cell(row=4,column=cc).fill=fill
    for j,h in enumerate(["Keyword","Vol","KD","Intent","YoY"]):
        cell=ws.cell(row=5,column=startcol+j,value=h); cell.font=Font(name="Arial",bold=True,size=9,color=htext)
        cell.fill=fill; cell.alignment=LEFT if j==0 else CTR; cell.border=BORDER
    rr=6
    for kw,vol,kd,intent,yoy,cpc in kws:
        ws.cell(row=rr,column=startcol,value=kw).font=TXT; ws.cell(row=rr,column=startcol).border=BORDER
        ws.cell(row=rr,column=startcol+1,value=vol if vol else "").font=TXT; ws.cell(row=rr,column=startcol+1).alignment=CTR; ws.cell(row=rr,column=startcol+1).border=BORDER
        ws.cell(row=rr,column=startcol+2,value=kd if kd is not None else "").font=TXT; ws.cell(row=rr,column=startcol+2).alignment=CTR; ws.cell(row=rr,column=startcol+2).border=BORDER
        ws.cell(row=rr,column=startcol+3,value=intent or "").font=TXT; ws.cell(row=rr,column=startcol+3).alignment=CTR; ws.cell(row=rr,column=startcol+3).border=BORDER
        ws.cell(row=rr,column=startcol+4,value=yoystr(yoy)).font=TXT; ws.cell(row=rr,column=startcol+4).alignment=CTR; ws.cell(row=rr,column=startcol+4).border=BORDER
        rr+=1
    return rr

def seg_block(ws, startcol, kws):
    # Segment tag + Vol + KD + Intent + YoY, sorted by segment then volume
    tagged=[(tag_keyword(kw),kw,vol,kd,intent,yoy) for kw,vol,kd,intent,yoy,cpc in kws]
    # order: by total segment volume desc, then kw volume desc; push Core/Unmodified last
    segvol={}
    for seg,kw,vol,*_ in tagged: segvol[seg]=segvol.get(seg,0)+(vol or 0)
    def sortkey(t):
        seg,kw,vol=t[0],t[1],(t[2] or 0)
        core = seg=="Core / Unmodified"
        return (1 if core else 0, -segvol.get(seg,0), seg, -vol)
    tagged.sort(key=sortkey)
    ws.merge_cells(start_row=4,start_column=startcol,end_row=4,end_column=startcol+4)
    tc=ws.cell(row=4,column=startcol,value="SEARCH SEGMENTS (auto-tagged · filter the Segment col)"); tc.font=Font(name="Arial",bold=True,size=10,color="041952")
    for cc in range(startcol,startcol+5): ws.cell(row=4,column=cc).fill=BLUESEG
    for j,h in enumerate(["Segment","Vol","KD","Intent","YoY"]):
        cell=ws.cell(row=5,column=startcol+j,value=h); cell.font=Font(name="Arial",bold=True,size=9,color="041952")
        cell.fill=BLUESEG; cell.alignment=LEFT if j==0 else CTR; cell.border=BORDER
    rr=6
    for seg,kw,vol,kd,intent,yoy in tagged:
        # show segment tag + the keyword it applies to (compact: 'Segment — keyword')
        ws.cell(row=rr,column=startcol,value=f"{seg}  ·  {kw}").font=TXT; ws.cell(row=rr,column=startcol).border=BORDER
        ws.cell(row=rr,column=startcol+1,value=vol if vol else "").font=TXT; ws.cell(row=rr,column=startcol+1).alignment=CTR; ws.cell(row=rr,column=startcol+1).border=BORDER
        ws.cell(row=rr,column=startcol+2,value=kd if kd is not None else "").font=TXT; ws.cell(row=rr,column=startcol+2).alignment=CTR; ws.cell(row=rr,column=startcol+2).border=BORDER
        ws.cell(row=rr,column=startcol+3,value=intent or "").font=TXT; ws.cell(row=rr,column=startcol+3).alignment=CTR; ws.cell(row=rr,column=startcol+3).border=BORDER
        ws.cell(row=rr,column=startcol+4,value=yoystr(yoy)).font=TXT; ws.cell(row=rr,column=startcol+4).alignment=CTR; ws.cell(row=rr,column=startcol+4).border=BORDER
        rr+=1
    return rr

for i,c in enumerate(order,1):
    gen=SPLIT[c]["generic"]; brd=SPLIT[c]["branded"]; d=ENT[c]
    gv=sum((x[1] or 0) for x in gen); bv=sum((x[1] or 0) for x in brd)
    ws=wb.create_sheet(tabname(i,c)); ws.sheet_view.showGridLines=False
    widths=[('A',2),('B',42),('C',9),('D',6),('E',11),('F',7),('G',2),
            ('H',42),('I',9),('J',6),('K',11),('L',7),('M',2),
            ('N',48),('O',9),('P',6),('Q',11),('R',7),('S',2),
            ('T',5),('U',32),('V',9),('W',2),
            ('X',34),('Y',13),('Z',11),('AA',40),('AB',2),
            ('AC',5),('AD',58),('AE',2),
            ('AF',5),('AG',48),('AH',46),('AI',2),
            ('AJ',24),('AK',20),('AL',24),('AM',24),('AN',2),
            ('AO',56),('AP',18),('AQ',16),('AR',2),
            ('AS',26),('AT',38),('AU',15),('AV',60)]
    for col,w in widths: ws.column_dimensions[col].width=w
    ws.merge_cells('B2:R2'); t=ws['B2']; t.value=f"CORE KEYWORD: {c.upper()}"; t.font=H1
    for col in range(2,19): ws.cell(row=2,column=col).fill=NAVY
    ws.row_dimensions[2].height=24
    ws.merge_cells('B3:R3'); s=ws['B3']
    s.value=f"GENERIC {len(gen)} (~{gv:,}/mo)  |  BRANDED {len(brd)} (~{bv:,}/mo)  |  segments auto-tagged  |  source: (US)"
    s.font=Font(name="Arial",bold=True,size=10,color="041952")
    for col in range(2,19): ws.cell(row=3,column=col).fill=LBLUE

    kw_block(ws,2,GREENHDR,"GENERIC KEYWORDS (unbranded)",gen,amber=True)
    kw_block(ws,8,ORANGE,"BRANDED KEYWORDS (competitor/retailer)",brd,amber=False)
    seg_block(ws,14,gen)

    # ---- TOP 10 COMPETITORS block (cols T-V = 20-22), before entity panel ----
    comps = COMP.get(c) or []
    ws.merge_cells(start_row=4,start_column=20,end_row=4,end_column=22)
    ch=ws.cell(row=4,column=20,value="TOP 10 ORGANIC COMPETITORS (for this core keyword)"); ch.font=Font(name="Arial",bold=True,size=10,color="041952")
    for col in range(20,23): ws.cell(row=4,column=col).fill=GREENHDR
    for j,h in enumerate(["#","Competitor Domain","Top Pos"]):
        cell=ws.cell(row=5,column=20+j,value=h); cell.font=Font(name="Arial",bold=True,size=9,color="041952")
        cell.fill=GREENHDR; cell.alignment=LEFT if j==1 else CTR; cell.border=BORDER
    cr=6
    for rank,(dom,pos,rating,etv) in enumerate(comps[:10],1):
        ws.cell(row=cr,column=20,value=rank).font=TXT; ws.cell(row=cr,column=20).alignment=CTR; ws.cell(row=cr,column=20).border=BORDER
        ws.cell(row=cr,column=21,value=dom).font=TXT; ws.cell(row=cr,column=21).border=BORDER
        ws.cell(row=cr,column=22,value=pos).font=TXT; ws.cell(row=cr,column=22).alignment=CTR; ws.cell(row=cr,column=22).border=BORDER
        cr+=1
    if not comps:
        ws.cell(row=6,column=21,value="(competitor pull pending)").font=NOTE; ws.cell(row=6,column=21).border=BORDER

    # ---- AI / AEO QUERIES block (cols X-AA = 24-27), after competitors ----
    aeo_rows = list(AEO.get(c) or [])
    # Prepend any live-harvested queries for this cluster's core so their real citations always show.
    import re as _re2
    def _ck2(s): return _re2.sub(r'[^a-z0-9 ]','',str(s).lower()).strip()
    existing_keys = { _ck2(r[0]) for r in aeo_rows }
    cluster_core = core_phrase(c).lower()
    live_prepend = []
    for lq, lv in (AEOLIVE or {}).items():
        # attach a live query to this cluster if its words relate to the core (share the core head noun)
        if _ck2(lq) in existing_keys: continue
        core_head = cluster_core.split()[-1] if cluster_core else ""
        if core_head and core_head in lq.lower():
            cited = ", ".join(lv.get("cited", [])) if lv.get("cited") else "pull live to capture"
            aio = "Yes (verified)" if lv.get("aio") else "No (verified)"
            live_prepend.append([lq, "", aio, cited])
            existing_keys.add(_ck2(lq))
    aeo_rows = live_prepend + aeo_rows
    ws.merge_cells(start_row=4,start_column=24,end_row=4,end_column=27)
    ah=ws.cell(row=4,column=24,value="AI / AEO QUERIES (AI Overview triggers + cited sources)"); ah.font=Font(name="Arial",bold=True,size=10,color="041952")
    for col in range(24,28): ws.cell(row=4,column=col).fill=AEOHDR
    for j,h in enumerate(["Conversational AI Query","Google Vol (proxy)","AI Overview?","Cited Sources (who AI quotes)"]):
        cell=ws.cell(row=5,column=24+j,value=h); cell.font=Font(name="Arial",bold=True,size=9,color="041952")
        cell.fill=AEOHDR; cell.alignment=LEFT if j in (0,3) else CTR; cell.border=BORDER
    ar=6
    import re as _re
    def _ck(s): return _re.sub(r'[^a-z0-9 ]','',str(s).lower()).strip()
    live_map = { _ck(k):v for k,v in (AEOLIVE or {}).items() }
    for q,vol,aio,cited in aeo_rows:
        lk = live_map.get(_ck(q))
        if lk:  # live-verified: override with real AIO flag + cited domains
            aio = "Yes (verified)" if lk.get("aio") else "No (verified)"
            if lk.get("cited"): cited = ", ".join(lk["cited"])
        ws.cell(row=ar,column=24,value=q).font=TXT; ws.cell(row=ar,column=24).alignment=Alignment(wrap_text=True,vertical="top"); ws.cell(row=ar,column=24).border=BORDER
        ws.cell(row=ar,column=25,value=vol if vol else "").font=TXT; ws.cell(row=ar,column=25).alignment=CTR; ws.cell(row=ar,column=25).border=BORDER
        ws.cell(row=ar,column=26,value=aio).font=TXT; ws.cell(row=ar,column=26).alignment=CTR; ws.cell(row=ar,column=26).border=BORDER
        ws.cell(row=ar,column=27,value=cited).font=TXT; ws.cell(row=ar,column=27).alignment=Alignment(wrap_text=True,vertical="top"); ws.cell(row=ar,column=27).border=BORDER
        ar+=1
    if not aeo_rows:
        ws.cell(row=6,column=24,value="(AI Overview check pending)").font=NOTE; ws.cell(row=6,column=24).border=BORDER
    else:
        fr=6+len(aeo_rows)+1
        ws.cell(row=fr,column=24,value="* Likely/Maybe = question-intent query not yet live-checked for AI Overview. 'Yes (verified)' + named sources = confirmed from live SERP. Per-LLM volume is not publicly available; Google Vol is a demand proxy.").font=NOTE
        ws.cell(row=fr,column=24).alignment=Alignment(wrap_text=True,vertical="top")
        ws.merge_cells(start_row=fr,start_column=24,end_row=fr,end_column=27)

    # ---- REDDIT URLs block (cols AC-AD = 29-30), top threads ranking across the cluster ----
    ryt = RYT.get(c) or {}
    reddit_urls = ryt.get("reddit", [])
    ws.merge_cells(start_row=4,start_column=29,end_row=4,end_column=30)
    rh=ws.cell(row=4,column=29,value="TOP REDDIT URLs (ranking across cluster keywords)"); rh.font=Font(name="Arial",bold=True,size=10,color="FFFFFF")
    for col in range(29,31): ws.cell(row=4,column=col).fill=REDDITHDR
    for j,h in enumerate(["#","Reddit Thread URL"]):
        cell=ws.cell(row=5,column=29+j,value=h); cell.font=Font(name="Arial",bold=True,size=9,color="FFFFFF")
        cell.fill=REDDITHDR; cell.alignment=LEFT if j==1 else CTR; cell.border=BORDER
    rr2=6
    for i,u in enumerate(reddit_urls[:20],1):
        ws.cell(row=rr2,column=29,value=i).font=TXT; ws.cell(row=rr2,column=29).alignment=CTR; ws.cell(row=rr2,column=29).border=BORDER
        ws.cell(row=rr2,column=30,value=u).font=TXT; ws.cell(row=rr2,column=30).alignment=Alignment(wrap_text=True,vertical="top"); ws.cell(row=rr2,column=30).border=BORDER
        rr2+=1
    if not reddit_urls:
        ws.cell(row=6,column=30,value="(Reddit harvest pending)").font=NOTE; ws.cell(row=6,column=30).border=BORDER

    # ---- YOUTUBE block (cols AF-AH = 32-34), videos ranking across the cluster ----
    yt = ryt.get("youtube", [])
    ws.merge_cells(start_row=4,start_column=32,end_row=4,end_column=34)
    yh=ws.cell(row=4,column=32,value="TOP YOUTUBE VIDEOS (ranking across cluster keywords)"); yh.font=Font(name="Arial",bold=True,size=10,color="FFFFFF")
    for col in range(32,35): ws.cell(row=4,column=col).fill=YTHDR
    for j,h in enumerate(["#","Video Title","URL"]):
        cell=ws.cell(row=5,column=32+j,value=h); cell.font=Font(name="Arial",bold=True,size=9,color="FFFFFF")
        cell.fill=YTHDR; cell.alignment=LEFT if j in (1,2) else CTR; cell.border=BORDER
    yr=6
    for i,item in enumerate(yt[:20],1):
        title,url = (item if isinstance(item,(list,tuple)) and len(item)==2 else ("",item))
        ws.cell(row=yr,column=32,value=i).font=TXT; ws.cell(row=yr,column=32).alignment=CTR; ws.cell(row=yr,column=32).border=BORDER
        ws.cell(row=yr,column=33,value=title).font=TXT; ws.cell(row=yr,column=33).alignment=Alignment(wrap_text=True,vertical="top"); ws.cell(row=yr,column=33).border=BORDER
        ws.cell(row=yr,column=34,value=url).font=TXT; ws.cell(row=yr,column=34).alignment=Alignment(wrap_text=True,vertical="top"); ws.cell(row=yr,column=34).border=BORDER
        yr+=1
    if not yt:
        ws.cell(row=6,column=33,value="(YouTube harvest pending)").font=NOTE; ws.cell(row=6,column=33).border=BORDER


    # entity panel W-Z (25-28)
    ws.merge_cells(start_row=4,start_column=36,end_row=4,end_column=39)
    eh=ws.cell(row=4,column=36,value="ENTITY PANEL (derived)"); eh.font=H2
    for col in range(36,40): ws.cell(row=4,column=col).fill=PURPLE
    for j,title in enumerate(["Synonyms","NLP Terms","Google Entities","AI Entities (derived)"]):
        darktxt = j<2
        cell=ws.cell(row=5,column=36+j,value=title); cell.font=Font(name="Arial",bold=True,size=9,color="041952" if darktxt else "FFFFFF")
        cell.fill=TEAL if j<2 else PURPLE; cell.alignment=CTR; cell.border=BORDER
    cols_data=[d["synonyms"],d["nlp"],d["google_entities"],d["ai_entities"]]
    maxlen=max((len(x) for x in cols_data),default=0)
    for k in range(maxlen):
        for j,lst in enumerate(cols_data):
            val=lst[k] if k<len(lst) else ""
            cell=ws.cell(row=6+k,column=36+j,value=val); cell.font=ENTF; cell.alignment=WRAP; cell.border=BORDER

    # ---- PAA BLOCK (cols Y-AA = 25-27), after entity panel ----
    pa = PAA.get(c) or {}
    paa_q = pa.get("paa", []); things = pa.get("things", []); related = pa.get("related", [])
    is_derived = pa.get("derived", False)
    core_vol = (gen[0][1] if gen and gen[0][1] else (brd[0][1] if brd and brd[0][1] else 1000))  # proxy: top keyword volume in cluster
    rows_paa = []
    # live PAA box: order = position (most prominent first); proxy rank = core_vol scaled by position
    for i,q in enumerate(paa_q):
        rows_paa.append((q, "PAA (derived)" if is_derived else "PAA (live)", core_vol - i))
    for i,q in enumerate(things):
        rows_paa.append((q, "Things to know", int(core_vol*0.6) - i))
    for i,q in enumerate(related):
        rows_paa.append((q, "Related question", int(core_vol*0.3) - i))
    # de-dupe by question text, keep highest proxy
    seen={}
    for q,t,pr in rows_paa:
        if q not in seen or pr>seen[q][2]: seen[q]=(q,t,pr)
    rows_paa=sorted(seen.values(), key=lambda r:-r[2])
    # header (cols AD-AF = 30-32)
    ws.merge_cells(start_row=4,start_column=41,end_row=4,end_column=43)
    ph=ws.cell(row=4,column=41,value="PEOPLE ALSO ASK / QUESTIONS (ranked by volume proxy)"); ph.font=Font(name="Arial",bold=True,size=10,color="041952")
    for col in range(41,44): ws.cell(row=4,column=col).fill=GREENHDR
    for j,h in enumerate(["Question","Type / Source","Vol proxy"]):
        cell=ws.cell(row=5,column=41+j,value=h); cell.font=Font(name="Arial",bold=True,size=9,color="041952")
        cell.fill=GREENHDR; cell.alignment=LEFT if j==0 else CTR; cell.border=BORDER
    pr=6
    for q,t,proxy in rows_paa:
        ws.cell(row=pr,column=41,value=q).font=TXT; ws.cell(row=pr,column=41).border=BORDER
        ws.cell(row=pr,column=41).alignment=Alignment(wrap_text=True,vertical="top")
        ws.cell(row=pr,column=42,value=t).font=TXT; ws.cell(row=pr,column=42).alignment=CTR; ws.cell(row=pr,column=42).border=BORDER
        ws.cell(row=pr,column=43,value=proxy if proxy>0 else "").font=TXT; ws.cell(row=pr,column=43).alignment=CTR; ws.cell(row=pr,column=43).border=BORDER
        pr+=1


    # ---- TOPICAL MAP BLOCK (cols AH-AK = 34-37), after PAA ----
    core = core_phrase(c)
    tmap = build_topical_map(core, n_sub=11, n_lt=10, n_titles=10)
    ws.merge_cells(start_row=4,start_column=45,end_row=4,end_column=48)
    th=ws.cell(row=4,column=45,value=f"TOPICAL MAP — {core} (click-worthy article titles)"); th.font=H2
    for col in range(45,49): ws.cell(row=4,column=col).fill=TOPHEAD
    for j,h in enumerate(["Sub-Cluster","Long-Tail Keyword","Intent","Click-Worthy Title"]):
        cell=ws.cell(row=5,column=45+j,value=h); cell.font=Font(name="Arial",bold=True,size=9,color="FFFFFF")
        cell.fill=TOPHEAD; cell.alignment=LEFT; cell.border=BORDER
    tr=6
    for sub,lt,intent,title in tmap:
        ws.cell(row=tr,column=45,value=sub).font=TXT; ws.cell(row=tr,column=45).alignment=Alignment(wrap_text=True,vertical="top"); ws.cell(row=tr,column=45).border=BORDER
        ws.cell(row=tr,column=46,value=lt).font=TXT; ws.cell(row=tr,column=46).alignment=Alignment(wrap_text=True,vertical="top"); ws.cell(row=tr,column=46).border=BORDER
        ws.cell(row=tr,column=47,value=intent).font=TXT; ws.cell(row=tr,column=47).alignment=CTR; ws.cell(row=tr,column=47).border=BORDER
        ws.cell(row=tr,column=48,value=title).font=TXT; ws.cell(row=tr,column=48).alignment=Alignment(wrap_text=True,vertical="top"); ws.cell(row=tr,column=48).border=BORDER
        tr+=1

    ws.freeze_panes="B6"

# remove README/active starter sheet so workbook opens straight into the data
for _name in list(wb.sheetnames):
    if _name.startswith("📋") or _name=="README" or _name=="Sheet":
        del wb[_name]
wb.save(OUT); print("Saved",OUT,"tabs",len(wb.worksheets))
